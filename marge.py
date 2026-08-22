import datetime
import io
import re
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st

st.set_page_config(page_title="Electricity Dept - Master Utility Portal", layout="wide")

# ==========================================
# SESSION STATE MANAGEMENT FOR MODE
# ==========================================
if "active_mode" not in st.session_state:
    st.session_state["active_mode"] = "URJAS"

# ==========================================
# SIDEBAR - NAVIGATION
# ==========================================
st.sidebar.title("🎛️ PORTAL NAVIGATION")
st.sidebar.markdown("---")

btn_urjas = st.sidebar.button(
    "⚡ URJAS PENDENCY PORTAL",
    use_container_width=True,
    type="primary" if st.session_state["active_mode"] == "URJAS" else "secondary"
)

st.sidebar.markdown("")

btn_merge = st.sidebar.button(
    "📊 MERGER & ZONE SPLITTER",
    use_container_width=True,
    type="primary" if st.session_state["active_mode"] == "MERGE" else "secondary"
)

if btn_urjas:
    st.session_state["active_mode"] = "URJAS"
    st.rerun()

if btn_merge:
    st.session_state["active_mode"] = "MERGE"
    st.rerun()

st.sidebar.markdown("---")
if st.session_state["active_mode"] == "URJAS":
    st.sidebar.success("📌 Active: **URJAS Master Portal**")
    st.sidebar.caption("👉 Create a Master Pendency & Time-Wise Dashboard by uploading a single Excel file.")
else:
    st.sidebar.success("📌 Active: **Universal Merger Portal**")
    st.sidebar.caption("👉 Merge multiple Excel files and make separate sheets for each zone.")


# ==============================================================================
# MODE 1: URJAS MASTER PENDENCY PORTAL
# ==============================================================================
if st.session_state["active_mode"] == "URJAS":
    st.title("⚡ Electricity Department - Master Pendency Report Generator")
    selected_date = st.date_input("📅 Target Pendency Date:", datetime.date.today())
    uploaded_file = st.file_uploader("Upload Raw Excel File (.xlsx)", type=["xlsx"])

    if uploaded_file:
        try:
            target_date = pd.to_datetime(selected_date).normalize()
            formatted_date_str = target_date.strftime("%d/%m/%Y")

            zones = [
                "ANNAPURNA",
                "GUMASTA NAGAR",
                "Hawa Bangla",
                "RAJ MOHALLA",
                "RAJENDRA NAGAR",
                "RAU",
                "SILICON CITY",
                "Sirpur",
            ]
            slabs = ["0 - 3 days", "4 - 6 days", "7 - 15 days", "16 - 30 days", "MORE THAN 30 DAYS"]

            # Broadened keywords and column patterns for seamless sheet matching
            sheet_configs = [
                {"title": "NSC LT Application", "keywords": ["nsc lt", "nsc", "new service"], "dc_patterns": ["DC", "ZONE"], "date_patterns": ["DATEOFAPPLICATION", "APP_DATE", "DATE"]},
                {"title": "LT Load Change", "keywords": ["load change", "load_change", "lt load"], "dc_patterns": ["DCNAME", "DC", "ZONE"], "date_patterns": ["DATEOFAPP", "DATE"]},
                {"title": "Meter Replacement App", "keywords": ["meter replacement", "meter_rep", "replacement"], "dc_patterns": ["DC", "ZONE"], "date_patterns": ["DATEOFAPPLICATION", "DATEOFAPP", "DATE"]},
                {"title": "Bill Correction App", "keywords": ["bill correction", "bill_corr", "correction"], "dc_patterns": ["DC", "ZONE"], "date_patterns": ["DATEOFAPP", "DATE"]},
                {"title": "Permanent Disconnection App", "keywords": ["permanent disconnection", "disconnection", "perm_disc"], "dc_patterns": ["DESCRIPTION", "DC", "ZONE"], "date_patterns": ["DATEOFAPP", "DATE"]},
                {"title": "LT Name Transfer App", "keywords": ["name transfer", "name_trans", "transfer"], "dc_patterns": ["DC", "ZONE"], "date_patterns": ["AADHARNO", "DATEOFAPP", "DATE"]},
                {"title": "LT Change Of Category", "keywords": ["change of category", "category_change", "category"], "dc_patterns": ["DC", "ZONE"], "date_patterns": ["DATEOFAPP", "DATE"]},
                {"title": "Cabel Replacement APP", "keywords": ["cabel", "cable", "cable_rep"], "dc_patterns": ["DC", "ZONE"], "date_patterns": ["DATEOFAPP", "DATE"]},
                {"title": "Transformer Fail App", "keywords": ["transformer", "transformer fail", "tf_fail"], "dc_patterns": ["DC", "ZONE"], "date_patterns": ["DATEOFAPPLICATION", "DATEOFAPP", "DATE"]},
                {"title": "LT Line/Meter Shifting App", "keywords": ["shifting", "meter shifting", "line shifting"], "dc_patterns": ["DC", "ZONE"], "date_patterns": ["DATEOFAPPLICATION", "DATEOFAPP", "DATE"]},
            ]

            xls = pd.ExcelFile(uploaded_file)
            available_sheets = xls.sheet_names

            wb = openpyxl.Workbook()

            font_title = Font(name="Arial", size=10, bold=True)
            font_header = Font(name="Arial", size=9, bold=True)
            font_body = Font(name="Arial", size=9)
            font_total = Font(name="Arial", size=9, bold=True)
            fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

            def clean_zone(z):
                if pd.isna(z): return "OTHER"
                z_str = str(z).strip().upper()
                if "ANNAPURNA" in z_str: return "ANNAPURNA"
                if "GUMASTA" in z_str: return "GUMASTA NAGAR"
                if "HAWA" in z_str or "BANGLA" in z_str: return "Hawa Bangla"
                if "MOHALLA" in z_str: return "RAJ MOHALLA"
                if "RAJENDRA" in z_str: return "RAJENDRA NAGAR"
                if "RAU" in z_str: return "RAU"
                if "SILICON" in z_str: return "SILICON CITY"
                if "SIRPUR" in z_str: return "Sirpur"
                return "OTHER"

            def get_slab(days):
                if pd.isna(days) or days < 0: return None
                try: days = int(days)
                except: return None
                if days <= 3: return "0 - 3 days"
                elif days <= 6: return "4 - 6 days"
                elif days <= 15: return "7 - 15 days"
                elif days <= 30: return "16 - 30 days"
                else: return "MORE THAN 30 DAYS"

            def parse_flexible_date(val):
                if pd.isna(val): return pd.NaT
                if isinstance(val, (datetime.datetime, datetime.date, pd.Timestamp)): return pd.to_datetime(val)
                val_str = str(val).replace("\xa0", "").strip()
                if val_str == "" or val_str.lower() in ["nan", "nat", "none"]: return pd.NaT
                try:
                    if val_str.replace(".", "", 1).isdigit():
                        num = float(val_str)
                        if num > 30000: return pd.to_datetime("1899-12-30") + pd.to_timedelta(num, unit="D")
                except: pass
                dt = pd.to_datetime(val_str, errors="coerce", dayfirst=True)
                if pd.isna(dt): dt = pd.to_datetime(val_str, errors="coerce", dayfirst=False)
                return dt

            def find_best_column(df, patterns):
                cols = list(df.columns)
                for pat in patterns:
                    for col in cols:
                        if pat.lower() in str(col).lower():
                            return col
                return cols[0] if cols else None

            ws_overall = wb.active
            ws_overall.title = "URJAS Pendency (Over all)"
            ws_overall.views.sheetView[0].showGridLines = True

            overall_summary = pd.DataFrame(0, index=zones, columns=[s["title"] for s in sheet_configs])
            processed_data_store = {}

            # Processing all sheets dynamically
            for config in sheet_configs:
                matched = next((s for s in available_sheets if any(k in s.lower() for k in config["keywords"])), None)
                if matched:
                    df = pd.read_excel(uploaded_file, sheet_name=matched)
                    zone_col = find_best_column(df, config["dc_patterns"])
                    date_col = find_best_column(df, config["date_patterns"])

                    if zone_col and date_col:
                        df["CleanZone"] = df[zone_col].apply(clean_zone)
                        app_dates = df[date_col].apply(parse_flexible_date)
                        df["Days"] = (target_date - app_dates).dt.days
                        df["Slab"] = df["Days"].apply(get_slab)
                        df = df[df["Days"] >= 0]

                        counts = df["CleanZone"].value_counts()
                        for z in zones:
                            overall_summary.at[z, config["title"]] = counts.get(z, 0)

                        processed_data_store[config["title"]] = df
                    else:
                        processed_data_store[config["title"]] = None
                else:
                    processed_data_store[config["title"]] = None

            # ----------------------------------------------------
            # SHEET 1: URJAS Pendency (Over all)
            # ----------------------------------------------------
            num_cols = len(sheet_configs) + 2
            ws_overall.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
            t_cell = ws_overall.cell(row=1, column=1, value=f"URJAS Pendency (Over all) Dtd.({formatted_date_str})")
            t_cell.fill = fill_yellow; t_cell.font = font_title
            t_cell.alignment = Alignment(horizontal="center", vertical="center")

            ws_overall.cell(row=2, column=1, value="Zone").fill = fill_yellow
            ws_overall.cell(row=2, column=1).font = font_header
            ws_overall.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws_overall.cell(row=2, column=1).border = thin_border

            for i, col_name in enumerate(overall_summary.columns):
                c = ws_overall.cell(row=2, column=i + 2, value=col_name)
                c.fill = fill_yellow; c.font = font_header
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border

            gt_head = ws_overall.cell(row=2, column=num_cols, value="Grand Total")
            gt_head.fill = fill_yellow; gt_head.font = font_header
            gt_head.alignment = Alignment(horizontal="center", vertical="center")
            gt_head.border = thin_border

            for r_idx, zone in enumerate(zones):
                row_num = r_idx + 3
                zn_cell = ws_overall.cell(row=row_num, column=1, value=zone)
                zn_cell.font = font_body; zn_cell.alignment = Alignment(horizontal="left", vertical="center"); zn_cell.border = thin_border

                row_tot = 0
                for c_idx, col_name in enumerate(overall_summary.columns):
                    val = int(overall_summary.at[zone, col_name])
                    val_cell = ws_overall.cell(row=row_num, column=c_idx + 2, value=val)
                    val_cell.font = font_body; val_cell.alignment = Alignment(horizontal="center", vertical="center"); val_cell.border = thin_border
                    row_tot += val

                tot_cell = ws_overall.cell(row=row_num, column=num_cols, value=row_tot)
                tot_cell.font = font_body; tot_cell.alignment = Alignment(horizontal="center", vertical="center"); tot_cell.border = thin_border

            tot_row_num = len(zones) + 3
            tot_label = ws_overall.cell(row=tot_row_num, column=1, value="Total")
            tot_label.fill = fill_yellow; tot_label.font = font_total; tot_label.alignment = Alignment(horizontal="center", vertical="center"); tot_label.border = thin_border

            for c_idx, col_name in enumerate(overall_summary.columns):
                col_sum = int(overall_summary[col_name].sum())
                col_sum_cell = ws_overall.cell(row=tot_row_num, column=c_idx + 2, value=col_sum)
                col_sum_cell.fill = fill_yellow; col_sum_cell.font = font_total; col_sum_cell.alignment = Alignment(horizontal="center", vertical="center"); col_sum_cell.border = thin_border

            grand_sum_val = int(overall_summary.values.sum())
            grand_sum_cell = ws_overall.cell(row=tot_row_num, column=num_cols, value=grand_sum_val)
            grand_sum_cell.fill = fill_yellow; grand_sum_cell.font = font_total; grand_sum_cell.alignment = Alignment(horizontal="center", vertical="center"); grand_sum_cell.border = thin_border

            # Adjusting Column Widths automatically
            for col in ws_overall.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws_overall.column_dimensions[col_letter].width = max(max_len + 3, 12)

            # ----------------------------------------------------
            # SHEET 2: Time Wise Pendency
            # ----------------------------------------------------
            ws_time = wb.create_sheet(title="Time Wise Pendency")
            ws_time.views.sheetView[0].showGridLines = True

            for idx, config in enumerate(sheet_configs):
                df = processed_data_store.get(config["title"])
                pvt = pd.DataFrame(0, index=zones, columns=slabs)

                if df is not None and not df.empty:
                    pvt_temp = pd.pivot_table(df, index="CleanZone", columns="Slab", aggfunc="size", fill_value=0)
                    pvt = pvt_temp.reindex(index=zones, columns=slabs, fill_value=0)

                col_side = idx % 2
                row_group = idx // 2
                start_col = 1 if col_side == 0 else 10
                start_row = 1 + (row_group * 13)

                ws_time.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 7)
                t_cell = ws_time.cell(row=start_row, column=start_col, value=f"{config['title'].upper()} TIME WISE PENDENCY TILL ({formatted_date_str})")
                t_cell.font = font_title; t_cell.fill = fill_yellow; t_cell.alignment = Alignment(horizontal="center", vertical="center")

                headers = ["SR/ NO.", "Zone", "0 - 3 days", "4 - 6 days", "7 - 15 days", "16 - 30 days", "MORE THAN 30 DAYS", "Grand Total"]
                h_row = start_row + 1
                for c_idx, h_text in enumerate(headers):
                    cell = ws_time.cell(row=h_row, column=start_col + c_idx, value=h_text)
                    cell.font = font_header; cell.fill = fill_yellow; cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = thin_border

                slab_totals = {s: 0 for s in slabs}
                cat_grand_total = 0

                for r_idx, z_name in enumerate(zones):
                    curr_row = h_row + 1 + r_idx
                    sr_cell = ws_time.cell(row=curr_row, column=start_col, value=r_idx + 1)
                    sr_cell.font = font_body; sr_cell.alignment = Alignment(horizontal="center", vertical="center"); sr_cell.border = thin_border
                    
                    zn_cell = ws_time.cell(row=curr_row, column=start_col + 1, value=z_name)
                    zn_cell.font = font_body; zn_cell.alignment = Alignment(horizontal="left", vertical="center"); zn_cell.border = thin_border

                    row_sum = 0
                    for s_idx, slab in enumerate(slabs):
                        val = int(pvt.loc[z_name, slab]) if z_name in pvt.index and slab in pvt.columns else 0
                        cell_val = ws_time.cell(row=curr_row, column=start_col + 2 + s_idx, value=val if val > 0 else "")
                        cell_val.font = font_body; cell_val.alignment = Alignment(horizontal="center", vertical="center"); cell_val.border = thin_border
                        row_sum += val
                        slab_totals[slab] += val

                    gt_cell = ws_time.cell(row=curr_row, column=start_col + 7, value=row_sum if row_sum > 0 else "")
                    gt_cell.font = font_body; gt_cell.alignment = Alignment(horizontal="center", vertical="center"); gt_cell.border = thin_border
                    cat_grand_total += row_sum

                tot_row = h_row + 1 + len(zones)
                ws_time.cell(row=tot_row, column=start_col, value="").border = thin_border
                
                lbl_tot = ws_time.cell(row=tot_row, column=start_col + 1, value="Total")
                lbl_tot.fill = fill_yellow; lbl_tot.font = font_total; lbl_tot.alignment = Alignment(horizontal="center", vertical="center"); lbl_tot.border = thin_border

                for s_idx, slab in enumerate(slabs):
                    s_tot = slab_totals[slab]
                    s_cell = ws_time.cell(row=tot_row, column=start_col + 2 + s_idx, value=s_tot if s_tot > 0 else 0)
                    s_cell.fill = fill_yellow; s_cell.font = font_total; s_cell.alignment = Alignment(horizontal="center", vertical="center"); s_cell.border = thin_border

                final_cat_gt = ws_time.cell(row=tot_row, column=start_col + 7, value=cat_grand_total if cat_grand_total > 0 else 0)
                final_cat_gt.fill = fill_yellow; final_cat_gt.font = font_total; final_cat_gt.alignment = Alignment(horizontal="center", vertical="center"); final_cat_gt.border = thin_border

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            st.success("🎉 URJAS Master Report successfully generated with ALL ZONES data!")
            st.download_button(
                label="📥 Download URJAS Master Complete Report (.xlsx)",
                data=output,
                file_name=f"URJAS_Master_Pendency_Report_{target_date.strftime('%d_%m_%Y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary"
            )

        except Exception as e:
            st.error(f"Error: {e}")


# ==============================================================================
# MODE 2: UNIVERSAL EXCEL MERGER & ZONE SPLITTER
# ==============================================================================
else:
    st.title("📊 Universal Excel Merger & Zone-Wise Data Generator")

    uploaded_files = st.file_uploader(
        "Upload Excel Files (.xlsx, .xls)",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
    )

    def process_clean_file(file, master_cols=None):
        xls = pd.ExcelFile(file)
        target_sheet = xls.sheet_names[0]
        for s in xls.sheet_names:
            if "data" in str(s).lower():
                target_sheet = s
                break

        df_raw = pd.read_excel(file, sheet_name=target_sheet, header=None)

        header_row_idx = 0
        for idx, row in df_raw.iterrows():
            row_vals = row.dropna().astype(str).str.strip().tolist()
            if len(row_vals) > 2:
                header_row_idx = idx
                break

        df = pd.read_excel(file, sheet_name=target_sheet, skiprows=header_row_idx)
        df = df.loc[:, ~df.columns.astype(str).str.contains("^Unnamed", na=False)]
        df.columns = [str(c).strip().upper() for c in df.columns]

        if not df.empty:
            first_col = df.columns[0]
            second_col = df.columns[1] if len(df.columns) > 1 else first_col

            df = df[~df[first_col].astype(str).str.upper().str.contains("GROUP NO|CONSUMER|S.NO|SL.NO", na=False)]
            df = df[~df[second_col].astype(str).str.upper().str.contains("GROUP NO|CONSUMER|NAME|ACCOUNT", na=False)]

            df = df[~df[first_col].astype(str).str.contains(r"TOTAL|GRAND TOTAL|RECORD", case=False, na=False)]
            df = df.dropna(how="all")

        if master_cols is not None:
            if len(df.columns) == len(master_cols):
                df.columns = master_cols
            else:
                df = df.reindex(columns=master_cols)

        return df

    if uploaded_files:
        combined_list = []
        master_headers = None

        for i, file in enumerate(uploaded_files):
            try:
                if i == 0:
                    cleaned_df = process_clean_file(file)
                    master_headers = list(cleaned_df.columns)
                else:
                    cleaned_df = process_clean_file(file, master_cols=master_headers)

                cleaned_df["SOURCE_FILE"] = file.name
                combined_list.append(cleaned_df)
            except Exception as e:
                st.error(f"⚠️ Error reading file {file.name}: {e}")

        if combined_list:
            merged_df = pd.concat(combined_list, ignore_index=True, axis=0)

            sno_col = next((c for c in merged_df.columns if "S.NO" in c or "SL.NO" in c or "S. NO" in c), None)
            if sno_col:
                merged_df[sno_col] = range(1, len(merged_df) + 1)

            st.success(f"✅ **Merged Successfully!** Total Records: **{len(merged_df):,}**")

            zone_col = next((c for c in merged_df.columns if "ZONE" in c or "DC" in c), None)

            col1, col2 = st.columns(2)

            with col1:
                st.markdown("### 📥 1. Download Master Merged File")
                output_single = io.BytesIO()
                with pd.ExcelWriter(output_single, engine="openpyxl") as writer:
                    merged_df.to_excel(writer, sheet_name="All_Data", index=False)

                st.download_button(
                    label="📄 Download All Data (.xlsx)",
                    data=output_single.getvalue(),
                    file_name="Master_Merged_All_Data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )

            with col2:
                st.markdown("### 📁 2. Download Zone-Wise Split File")
                if zone_col:
                    output_zone = io.BytesIO()
                    with pd.ExcelWriter(output_zone, engine="openpyxl") as writer:
                        merged_df.to_excel(writer, sheet_name="All_Data", index=False)
                        merged_df[zone_col] = merged_df[zone_col].astype(str).str.strip()
                        unique_zones = [z for z in merged_df[zone_col].unique() if z and z.lower() != "nan"]

                        for z in sorted(unique_zones):
                            z_df = merged_df[merged_df[zone_col] == z].copy()
                            if not z_df.empty:
                                if sno_col in z_df.columns:
                                    z_df[sno_col] = range(1, len(z_df) + 1)
                                sheet_name = re.sub(r"[:*?/\\[\]]", "_", str(z))[:30]
                                z_df.to_excel(writer, sheet_name=sheet_name, index=False)

                    st.download_button(
                        label="📊 Download Zone-Wise Excel (.xlsx)",
                        data=output_zone.getvalue(),
                        file_name="Zone_Wise_Separated.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                    )
                else:
                    st.warning("⚠️ Zone Column nahi mila.")

            st.markdown("---")
            st.markdown("### 📋 Data Preview")
            st.dataframe(merged_df, use_container_width=True)
